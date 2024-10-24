import openpyxl
import os
from graphs import graph_define
from langdetect import detect
from collections import Counter

LABLES = ['Зам. директора', 'Заместитель директора, учитель физкультуры, биологии', 'Учитель английского языка',
     'учитель английского языка', 'учитель начальных классов', 'Учитель географии',
     'Социальный педагог', 'Заместитель директора', 'Учитель - логопед', 'Учитель химии, биологии',
     'Учитель начальных классов', 'Заместитель директора  по УД', 'Учитель английского языка',
     'Директор', 'Учитель  истории и обществознания', 'Учитель начальных классов',
     'Учитель русского языка и литературы', 'Учитель', 'Учитель физики', 'Учитель русского языка, английского языка',
     'Учитель математики', 'Учитель музыки и изо', 'Учитель начальных классов', 'Учитель - логопед',
     'Заместитель директора, учитель ОБЖ', 'Учитель русского языка и литературы', 'Учитель информатики',
     'Учитель математики', 'Советник по воспитанию, учитель истории', 'Педагог – организатор ,', 'Учитель музыки',
     'Заместитель директора, учитель истории и обществознания', 'Учитель начальных классов',
     'Учитель математики', 'Учитель русского языка и литературы', 'Учитель технологии', 'Учитель географии',
     'Учитель физической культуры', 'Учитель начальных классов', 'Учитель технологии', 'Учитель математики',
     'Учитель начальных классов', 'Учитель английского языка']

SIMILAR_LETTERS = {
    'a':'а',
    'b':'ь',
    'c':'с',
    'e':'е',
    'k':'к',
    'm':'м',
    'n':'п',
    'o':'о',
    'p':'р',
    'r':'г',
    'u':'и',
    'x':'х',
    'y':'у',
    'A':'А',
    'B':'В',
    'C':'С',
    'E':'Е',
    'H':'Н',
    'K':'К',
    'M':'М',
    'O':'О',
    'P':'Р',
    'T':'Т',
    'X':'Х',
    'Y':'У'
}
USERNAME = os.getlogin()
BASE_DIR = f'C:/Users/{USERNAME}/Documents/Social_capital'
CACHE_FILE = f'C:/Users/{USERNAME}/Documents/Social_capital/CACHE.xlsx'
# Необходимо заменить значение переменной REPORT_NAME на имя анализируемого результата анкетирования
REPORT_NAME = '2024-10-12 MAOU SOSh 87'
WORKBOOK = openpyxl.open(f'{BASE_DIR}/{REPORT_NAME}.xlsx')
WORKSHEET = WORKBOOK['Sheet']
FIRST_MATRIX_QUESTION = '«Кто из коллег, по вашему мнению, являются лучшими педагогами (преподавателями, воспитателями) вашей образовательной организации?»'

def create_cache(filename):
    """Функция создания кэша, где хранятся имена отчетов, уже проанализированных"""
    cache = openpyxl.Workbook(filename)
    cache.save(filename)
    return


def check_cache(filename, reportname):
    """Функция проверки записи в кэше, если отчет уже подвергался анализу, повторное выполнение не будет выполнено."""
    cachebook = openpyxl.open(filename)
    cachesheet = cachebook['Sheet']
    rows = cachesheet.max_row
    for current_row in range(1, rows+1):
        if cachesheet.cell(row = current_row, column=1).value == reportname:
            print('the record of the report already exists')
            return
    cachesheet.cell(row = rows, column=1).value = reportname
    cachebook.save(filename)
    print('A new entry has been added')
    return


def get_teachers_list(WORKSHEET):
    """Создание списка преподавателей"""
    teachers_list = []
    rows = WORKSHEET.max_row
    for current_row in range(2, rows+1):
        if WORKSHEET.cell(row=current_row, column=1).value not in teachers_list:
            teachers_list.append(WORKSHEET.cell(row=current_row, column=1).value)
        else:
            pass
    return sorted(list(set(teachers_list)))


def teatchers_dict(teachers_list):
    """Создание словаря для вычисления индексов преподавателей в матрице"""
    teachers_dict = {}
    for index in range(len(teachers_list)):
        teachers_dict[teachers_list[index]]=index+2
    return teachers_dict


def create_matrix(teachers, amount, workbook):
    """Создание матрицы выбора лучшего преподавателя"""
    workbook.create_sheet('matrix_best_teacher')
    worksheet = workbook['matrix_best_teacher']
    for current_cell in range(2, amount+2):
        worksheet.cell(row=current_cell, column=1).value = teachers[current_cell-2]
        worksheet.cell(row=1, column=current_cell).value = teachers[current_cell-2]
    workbook.save(f'{BASE_DIR}/{REPORT_NAME}.xlsx')
    return worksheet


def first_matrix_start(question, data):
    """Поиск начала первой матрицы по вопросу"""
    columns = data.max_column
    for current_column in range(1, columns+1):
        if question in data.cell(row=1, column=current_column).value:
            return current_column
        else:
            pass


def symbol_change(text):
    """Замена латинских символов на кириллицу"""
    result = ''
    for letter in text:
        if letter == ' ':
            result += letter
        elif detect(letter) != 'en':
            result += letter
        else:
            result += SIMILAR_LETTERS[letter]
    return result


def fill_the_matrix(matrix, start, data, amount, workbook, teachers_indexes):
    """Заполнение матрицы данными"""
    for current_row in range(2, amount+2):
        for current_column in range(start, start+amount):
            if data.cell(row=current_row, column=current_column).value is not None:
                if data.cell(row=current_row, column=current_column).value in teachers_indexes:
                    matrix_row = teachers_indexes[data.cell(row=current_row, column=1).value]
                    matrix_column = teachers_indexes[data.cell(row=current_row, column=current_column).value]
                    matrix.cell(row=matrix_row, column=matrix_column).value = 1
                else:
                    name = symbol_change(data.cell(row=current_row, column=1).value)
                    matrix_row = teachers_indexes[data.cell(row=current_row, column=1).value]
                    matrix_column = teachers_indexes[name]
                    matrix.cell(row=matrix_row, column=matrix_column).value = 1
            else:
                if data.cell(row=1, column=current_column).value[133:] in teachers_indexes:
                    matrix_row = teachers_indexes[data.cell(row=current_row, column=1).value]
                    matrix_column = teachers_indexes[data.cell(row=1, column=current_column).value[133:]]
                    matrix.cell(row=matrix_row, column=matrix_column).value = 0
                else:
                    name = symbol_change(data.cell(row=1, column=current_column).value[133:])
                    matrix_row = teachers_indexes[data.cell(row=current_row, column=1).value]
                    matrix_column = teachers_indexes[name]
                    matrix.cell(row=matrix_row, column=matrix_column).value = 0
    workbook.save(f'{BASE_DIR}/{REPORT_NAME}.xlsx')
    return workbook['matrix_best_teacher']


def get_nodes(teachers_indexes):
    """Создание списка узлов"""
    return [teacher[1] for teacher in teachers_indexes.items()]


def get_two_side_nodes(two_side_edges):
    """Создание узлов с обратной связью"""
    return list(set([edge[0] for edge in two_side_edges]))


def grt_three_side_nodes():
    pass

def get_edges(matrix, teachers_indexes, amount, workbook):
    """Создание массива ребер для построения графа"""
    edges = []
    workbook.create_sheet('edges')
    worksheet = WORKBOOK['edges']
    start_row = 1
    for current_row in range(2, amount+3):
        for current_column in range(2, amount+3):
            if matrix.cell(row=current_row, column=current_column).value == 1:
                edges.append(
                    (
                        teachers_indexes[matrix.cell(row=current_row, column=1).value],
                        teachers_indexes[matrix.cell(row=1, column=current_column).value]
                    )
                )
                worksheet.cell(row=start_row, column=1).value = teachers_indexes[matrix.cell(row=current_row, column=1).value]
                worksheet.cell(row=start_row, column=2).value = teachers_indexes[matrix.cell(row=1, column=current_column).value]
                start_row += 1
            else:
                pass
    workbook.save(f'{BASE_DIR}/{REPORT_NAME}.xlsx')
    return edges


def get_lables(nodes):
    """Получение должностей в зависимости от списка узлов"""
    lables = []
    for node in nodes:
        lables.append(LABLES[node-2])
    return lables


def get_two_side_edges(edges):
    """Получение ребер с обратной связью"""
    two_side_edges = []
    for index in range(len(edges)-1):
        for second_index in range(1, len(edges)):
            if edges[index] == (edges[second_index][1], edges[second_index][0]):
                two_side_edges.append(edges[index])
                two_side_edges.append(edges[second_index])
            else:
                pass
    return two_side_edges


def get_three_side_nodes(two_side_edges):
    """Получение ребер взаимных связей между 3 преподавателями"""
    result = []
    for index in range(0, len(two_side_edges) - 4, 2):
        temp = two_side_edges[index]
        for second_index in range(index + 2, len(two_side_edges) - 2, 2):
            if temp[0] in two_side_edges[second_index]:
                temp_2 = list(two_side_edges[second_index])
                temp_2.remove(temp[0])
                if (temp[1], temp_2[0]) in two_side_edges[second_index + 2:]:
                    result.append((two_side_edges[index], two_side_edges[index + 1],
                                   two_side_edges[second_index], two_side_edges[second_index + 1],
                                   (temp[1], temp_2[0]), (temp_2[0], temp[1])))
    return result



def get_node_size(edges, nodes):
    """Задание массы узла в зависимости от количества связей"""
    sizes = {}
    for node in nodes:
        node -= 1
        sizes[node] = 5
    for edge in edges:
        node_number = edge[1]-1
        if node_number in sizes:
            sizes[node_number] += 1
        else:
            pass
    return sizes


def get_titles(teachers_indexes):
    """Получение заголовков для построения графа"""
    return list(teachers_indexes.keys())


def get_two_side_titles(nodes, teachers_indexes):
    """Получение заголовка для двухсторонних связей"""
    titles = []
    for title, id in teachers_indexes.items():
        if id in nodes:
            titles.append(title)
    return titles


if __name__ == '__main__':
    if os.path.exists(BASE_DIR):
        if os.path.exists(CACHE_FILE):
            print('the cache file exists')
            pass
        else:
            print('the cache file has been created')
            create_cache(CACHE_FILE)
    else:
        os.mkdir(BASE_DIR)
        create_cache(CACHE_FILE)
        print('the directory has been created, the cache file has been created')
    # check_cache(CACHE_FILE, REPORT_NAME)
    teachers_list = get_teachers_list(WORKSHEET)
    teachers_amount = len(teachers_list)
    teachers_indexes = teatchers_dict(teachers_list)
    matrix = create_matrix(teachers_list, teachers_amount, WORKBOOK)
    first_matrix_pivot = first_matrix_start(FIRST_MATRIX_QUESTION, WORKSHEET)
    filled_matrix = fill_the_matrix(matrix, first_matrix_pivot, WORKSHEET, teachers_amount, WORKBOOK, teachers_indexes)
    edges = get_edges(filled_matrix, teachers_indexes, teachers_amount, WORKBOOK)
    nodes = get_nodes(teachers_indexes)
    nodes_sizes = get_node_size(edges, nodes)
    titles = get_titles(teachers_indexes)
    lables = get_lables(nodes)
    # graph_define(nodes, edges, titles, nodes_sizes, lables)
    two_side_edges = get_two_side_edges(edges)
    two_side_nodes = get_two_side_nodes(two_side_edges)
    two_side_nodes_size = get_node_size(two_side_edges, two_side_nodes)
    two_side_titles = get_two_side_titles(two_side_nodes, teachers_indexes)
    two_side_lables = get_lables(two_side_nodes)
    # graph_define(two_side_nodes, two_side_edges, two_side_titles, two_side_nodes_size, two_side_lables)
    print(two_side_edges)
